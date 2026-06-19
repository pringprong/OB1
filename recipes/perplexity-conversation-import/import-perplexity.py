#!/usr/bin/env python3
"""
Open Brain — Perplexity Export Importer

Extracts conversations and memories from a Perplexity data export (.xlsx),
filters deleted/already-imported items, summarizes conversations into
1-3 distilled thoughts via LLM, and loads everything into your Open Brain.

Memory entries are ingested directly (already summarized by Perplexity).
JSON profile rows (MEMORY_KEY empty, MEMORY_VALUE is a JSON object) are
flattened into per-section thoughts.

Usage:
    python import-perplexity.py path/to/export.xlsx [options]

Ingestion modes:
    Default:              Supabase direct insert (requires SUPABASE_URL,
                          SUPABASE_SERVICE_ROLE_KEY, OPENROUTER_API_KEY)

Options:
    --xlsx PATH           Path to Perplexity .xlsx export (required)
    --dry-run             Parse, filter, summarize, but don't ingest
    --after YYYY-MM-DD    Only conversations created after this date
    --before YYYY-MM-DD   Only conversations created before this date
    --limit N             Max items per type to process
    --type TYPE           What to import: conversations, memory, or both (default: both)
    --model MODEL         LLM backend: openrouter (default) or ollama
    --ollama-model NAME   Ollama model name (default: qwen3)
    --verbose             Show full content during processing
    --report FILE         Write a markdown report of everything imported

Environment variables:
    SUPABASE_URL               Supabase project URL
    SUPABASE_SERVICE_ROLE_KEY  Supabase service role key
    OPENROUTER_API_KEY         OpenRouter API key (summarization + embeddings)
"""

import argparse
import hashlib
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# ─── Configuration ───────────────────────────────────────────────────────────

SYNC_LOG_PATH = Path("perplexity-sync-log.json")

OPENROUTER_BASE = "https://openrouter.ai/api/v1"
OLLAMA_BASE = "http://localhost:11434"

# Supabase: reads the "Secret Key" from Settings → API (starts with sb_secret_)
# The env var name uses the legacy convention for cross-recipe consistency.
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")

SUMMARIZATION_PROMPT = """\
You are distilling a Perplexity Q&A exchange into standalone thoughts for a \
personal knowledge base. Your job is to be HIGHLY SELECTIVE — only extract \
knowledge that would be valuable to retrieve months or years from now.

You will receive a search query and Perplexity's answer.

CAPTURE these (1-3 thoughts max):
- Decisions made and the reasoning behind them
- People, places, or topics explored with lasting relevance
- Lessons learned, preferences discovered, or useful frameworks
- Research findings worth remembering
- Context about the user's interests, projects, or goals

SKIP these entirely (return empty):
- Simple factual lookups (restaurant hours, definitions, recipes)
- One-off trivia with no lasting value
- Generic how-to with no personal context

Each thought must be:
- A clear, standalone statement (makes sense without the Q&A)
- Written in first person
- Anchored with names, dates, or context when available
- 1-3 sentences

Return JSON: {"thoughts": ["thought1", "thought2"]}
If nothing worth capturing, return {"thoughts": []}
Err on the side of returning empty — less is more."""

# JSON profile flattening config
PROFILE_SECTIONS = {
    "demographics": "Demographics",
    "interests": "Interests",
    "work_and_education": "Work and Education",
    "lifestyle": "Lifestyle",
    "technology": "Technology",
    "knowledge": "Knowledge and Expertise",
    "personal_traits": "Personal Traits",
}

# ─── Sync Log ────────────────────────────────────────────────────────────────


def load_sync_log():
    """Load sync log from disk. Returns dict with ingested_ids and last_sync."""
    try:
        with open(SYNC_LOG_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"ingested_ids": {}, "last_sync": ""}


def save_sync_log(log):
    """Save sync log to disk."""
    with open(SYNC_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2)


def make_dedupe_key(*parts):
    """Generate a short SHA256 hash from string parts for deduplication."""
    raw = "|".join(str(p) for p in parts)
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


# ─── HTTP Helpers ────────────────────────────────────────────────────────────

try:
    import requests
except ImportError:
    print("Missing dependency: requests")
    print("Install with: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def http_post_with_retry(url, headers, body, retries=2):
    """POST with exponential backoff retry on transient failures."""
    for attempt in range(retries + 1):
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=30)
            if resp.status_code >= 500 and attempt < retries:
                time.sleep(1 * (attempt + 1))
                continue
            return resp
        except requests.RequestException:
            if attempt < retries:
                time.sleep(1 * (attempt + 1))
                continue
            raise
    return None  # unreachable


# ─── XLSX Parsing ────────────────────────────────────────────────────────────


def _cell_value(cell):
    """Extract a clean string value from a cell, treating None as empty."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


TIMESTAMP_FORMATS = (
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
)


def parse_timestamp_iso(raw):
    """Parse a Perplexity timestamp string into ISO 8601 (UTC).

    Returns ISO string or None if parsing fails.
    Also handles datetime objects from openpyxl (data_only mode).
    """
    if not raw:
        return None

    # openpyxl may return datetime objects directly
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%dT%H:%M:%S+00:00")

    raw = str(raw).strip()
    if not raw:
        return None

    for fmt in TIMESTAMP_FORMATS:
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%Y-%m-%dT%H:%M:%S+00:00")
        except ValueError:
            continue

    return None


def extract_conversations(xlsx_path):
    """Extract conversation rows from the 'Conversations' sheet.

    Returns list of dicts with keys: uuid, created, updated, title, answer_text.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Conversations" not in wb.sheetnames:
        print("Warning: No 'Conversations' sheet found in export.")
        wb.close()
        return []

    ws = wb["Conversations"]
    rows = list(ws.iter_rows())
    wb.close()

    if len(rows) < 2:
        return []

    header = [_cell_value(c) for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(header)}

    conversations = []
    for row in rows[1:]:
        values = [_cell_value(c) for c in row]

        uuid = values[col_idx.get("UUID", -1)] if "UUID" in col_idx else ""
        created = values[col_idx.get("CREATED", -1)] if "CREATED" in col_idx else ""
        updated = values[col_idx.get("UPDATED", -1)] if "UPDATED" in col_idx else ""
        title = values[col_idx.get("TITLE", -1)] if "TITLE" in col_idx else ""
        output_str = (
            values[col_idx.get("OUTPUT_STR", -1)] if "OUTPUT_STR" in col_idx else ""
        )

        answer_text = _parse_output_str(output_str)

        if not uuid and not title:
            continue

        conversations.append(
            {
                "uuid": uuid,
                "created": created,
                "updated": updated,
                "title": title,
                "answer_text": answer_text,
            }
        )

    return conversations


def _parse_output_str(output_str):
    """Extract answer text from Perplexity's OUTPUT_STR JSON blob."""
    if not output_str:
        return ""

    try:
        data = json.loads(output_str)
    except (json.JSONDecodeError, TypeError):
        return output_str

    if isinstance(data, dict):
        answer = data.get("answer", "")
        if isinstance(answer, str):
            return answer.strip()
        return json.dumps(answer) if answer else ""

    return output_str


def extract_memory_rows(xlsx_path):
    """Extract memory rows from the 'Memory' sheet.

    Returns list of dicts with all memory columns.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Memory" not in wb.sheetnames:
        print("Warning: No 'Memory' sheet found in export.")
        wb.close()
        return []

    ws = wb["Memory"]
    rows = list(ws.iter_rows())
    wb.close()

    if len(rows) < 2:
        return []

    header = [_cell_value(c) for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(header)}

    memories = []
    for row in rows[1:]:
        values = [_cell_value(c) for c in row]

        mem = {}
        for col_name, idx in col_idx.items():
            mem[col_name] = values[idx] if idx < len(values) else ""

        # Normalize boolean-ish strings
        for bool_col in ("IS_DELETED", "IS_FORGOTTEN", "IS_INVISIBLE"):
            val = mem.get(bool_col, "").lower()
            mem[bool_col] = val in ("true", "1", "yes")

        if not mem.get("MEMORY_KEY") and not mem.get("MEMORY_VALUE"):
            continue

        memories.append(mem)

    return memories


# ─── JSON Profile Handling ──────────────────────────────────────────────────


def is_json_profile_row(row):
    """Detect if a memory row is a JSON profile (no MEMORY_KEY, MEMORY_VALUE is JSON)."""
    memory_key = row.get("MEMORY_KEY", "").strip()
    memory_value = row.get("MEMORY_VALUE", "").strip()

    if memory_key:
        return False

    if not memory_value:
        return False

    if not memory_value.startswith("{"):
        return False

    try:
        json.loads(memory_value)
        return True
    except (json.JSONDecodeError, TypeError):
        return False


def flatten_json_section(key, value):
    """Flatten a JSON section into a natural-language string.

    Example input:
        key="demographics", value={"languages": ["Swedish", "English"], "locations": ["Sweden"]}

    Example output:
        "Languages: Swedish, English. Locations: Sweden."
    """
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    if not isinstance(value, dict):
        return json.dumps(value)

    parts = []
    for sub_key, sub_value in value.items():
        label = sub_key.replace("_", " ").title()
        if isinstance(sub_value, list):
            items = [str(v) for v in sub_value if v]
            if items:
                parts.append(f"{label}: {', '.join(items)}")
        elif isinstance(sub_value, str) and sub_value.strip():
            parts.append(f"{label}: {sub_value.strip()}")
        elif isinstance(sub_value, dict):
            nested = flatten_json_section(sub_key, sub_value)
            if nested:
                parts.append(f"{label}: {nested}")

    return ". ".join(parts) + "." if parts else ""


def flatten_json_profile(json_obj):
    """Flatten a JSON profile into a list of (synthetic_key, text) tuples.

    The 'summary' field becomes one thought, each recognized section becomes another.
    """
    results = []

    # Summary first
    summary = json_obj.get("summary", "")
    if isinstance(summary, str) and summary.strip():
        results.append(("profile.summary", summary.strip()))

    # Flattened sections
    for section_key, label in PROFILE_SECTIONS.items():
        section_data = json_obj.get(section_key)
        if section_data is None:
            continue
        text = flatten_json_section(section_key, section_data)
        if text:
            results.append((f"profile.{section_key}", text))

    return results


# ─── Filtering ───────────────────────────────────────────────────────────────


def should_skip_conversation(conv, sync_log, args):
    """Return a skip reason string, or None if the conversation should be processed."""
    dedupe_key = make_dedupe_key(conv["uuid"])

    if dedupe_key in sync_log["ingested_ids"]:
        return "already_imported"

    # Date filtering on CREATED
    created = conv.get("created", "")
    if created:
        try:
            # Handle both "2023-12-24 10:00:00" and "2023-12-24T10:00:00" formats
            for fmt in (
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d",
            ):
                try:
                    conv_date = datetime.strptime(created, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                conv_date = None

            if conv_date:
                if args.after and conv_date < args.after:
                    return "before_date_filter"
                if args.before and conv_date > args.before:
                    return "after_date_filter"
        except Exception:
            pass

    return None


def should_skip_memory(row, sync_log):
    """Return a skip reason string, or None if the memory row should be processed."""
    # Skip deleted or forgotten
    if row.get("IS_DELETED"):
        return "deleted"
    if row.get("IS_FORGOTTEN"):
        return "forgotten"

    # Deduplication key: UUID-based for JSON profiles, key-based for normal rows
    if is_json_profile_row(row):
        dedupe_key = make_dedupe_key("json_profile", row.get("MEMORY_VALUE", "")[:200])
    else:
        dedupe_key = make_dedupe_key(
            row.get("MEMORY_KEY", ""), row.get("FIRST_CREATED_AT", "")
        )

    if dedupe_key in sync_log["ingested_ids"]:
        return "already_imported"

    return None


# ─── LLM Summarization ──────────────────────────────────────────────────────


def summarize_openrouter(title, date_str, answer_text):
    """Summarize a Perplexity Q&A into thoughts using OpenRouter."""
    if not OPENROUTER_API_KEY:
        print("   Warning: Skipping summarization (no OPENROUTER_API_KEY)")
        return []

    truncated = answer_text[:6000]

    resp = http_post_with_retry(
        f"{OPENROUTER_BASE}/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        body={
            "model": "deepseek/deepseek-v4-flash",
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": SUMMARIZATION_PROMPT},
                {
                    "role": "user",
                    "content": f"Search query: {title}\nDate: {date_str}\n\nPerplexity's answer:\n{truncated}",
                },
            ],
            "temperature": 0,
        },
    )

    if not resp or resp.status_code != 200:
        status = resp.status_code if resp else "no response"
        print(f"   Warning: Summarization failed ({status}), skipping conversation.")
        return []

    try:
        data = resp.json()
        result = json.loads(data["choices"][0]["message"]["content"])
        thoughts = result.get("thoughts", [])
        return [t for t in thoughts if isinstance(t, str) and t.strip()]
    except (KeyError, json.JSONDecodeError, IndexError) as e:
        print(f"   Warning: Failed to parse summarization response: {e}")
        return []


def summarize_ollama(title, date_str, answer_text, model_name="qwen3"):
    """Summarize a Perplexity Q&A using a local Ollama model."""
    truncated = answer_text[:6000]

    prompt = (
        f"{SUMMARIZATION_PROMPT}\n\n"
        f"Search query: {title}\nDate: {date_str}\n\n"
        f"Perplexity's answer:\n{truncated}"
    )

    try:
        resp = requests.post(
            f"{OLLAMA_BASE}/api/generate",
            json={
                "model": model_name,
                "prompt": prompt,
                "stream": False,
                "format": "json",
            },
            timeout=120,
        )
    except requests.RequestException as e:
        print(f"   Warning: Ollama request failed: {e}")
        return []

    if resp.status_code != 200:
        print(f"   Warning: Ollama returned {resp.status_code}")
        return []

    try:
        raw = resp.json().get("response", "")
        result = json.loads(raw)
        thoughts = result.get("thoughts", [])
        return [t for t in thoughts if isinstance(t, str) and t.strip()]
    except (json.JSONDecodeError, KeyError) as e:
        print(f"   Warning: Failed to parse Ollama response: {e}")
        return []


def summarize(title, date_str, answer_text, args):
    """Dispatch to the appropriate summarization backend."""
    if args.model == "ollama":
        return summarize_ollama(title, date_str, answer_text, args.ollama_model)
    return summarize_openrouter(title, date_str, answer_text)


# ─── Embedding Generation ───────────────────────────────────────────────────


def generate_embedding(text):
    """Generate a 1024-dim embedding via OpenRouter (text-embedding-3-small)."""
    truncated = text[:8000]

    resp = http_post_with_retry(
        f"{OPENROUTER_BASE}/embeddings",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        body={
            "model": "intfloat/multilingual-e5-large",
            "input": truncated,
        },
    )

    if not resp or resp.status_code != 200:
        status = resp.status_code if resp else "no response"
        print(f"   Warning: Embedding generation failed ({status})")
        return None

    try:
        data = resp.json()
        return data["data"][0]["embedding"]
    except (KeyError, IndexError) as e:
        print(f"   Warning: Failed to parse embedding response: {e}")
        return None


# ─── Ingestion ───────────────────────────────────────────────────────────────


def ingest_thought_supabase(content, metadata_dict, created_at=None):
    """Insert a thought directly into Supabase with a generated embedding."""
    embedding = generate_embedding(content)
    if not embedding:
        return {"ok": False, "error": "Failed to generate embedding"}

    body = {
        "content": content,
        "embedding": embedding,
        "metadata": metadata_dict,
    }
    if created_at:
        body["created_at"] = created_at

    resp = http_post_with_retry(
        f"{SUPABASE_URL}/rest/v1/thoughts",
        headers={
            "Content-Type": "application/json",
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Prefer": "return=minimal",
        },
        body=body,
    )

    if not resp:
        return {"ok": False, "error": "No response from Supabase"}

    if resp.status_code not in (200, 201):
        try:
            error_detail = resp.json()
        except ValueError:
            error_detail = resp.text
        return {"ok": False, "error": f"HTTP {resp.status_code}: {error_detail}"}

    return {"ok": True}


# ─── CLI ─────────────────────────────────────────────────────────────────────


def parse_date(s):
    """Parse a YYYY-MM-DD string to a date object."""
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        print(f"Error: Invalid date format '{s}'. Use YYYY-MM-DD.")
        sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import Perplexity conversations and memories into Open Brain",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  python import-perplexity.py export.xlsx --dry-run --limit 5
  python import-perplexity.py export.xlsx --type memory
  python import-perplexity.py export.xlsx --type conversations --after 2024-01-01
  python import-perplexity.py export.xlsx --model ollama --ollama-model qwen3
  python import-perplexity.py export.xlsx --report import-report.md""",
    )
    parser.add_argument("xlsx_path", help="Path to Perplexity data export .xlsx file")
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse and summarize but don't ingest"
    )
    parser.add_argument(
        "--after", type=parse_date, help="Only conversations after YYYY-MM-DD"
    )
    parser.add_argument(
        "--before", type=parse_date, help="Only conversations before YYYY-MM-DD"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Max items per type to process (0 = unlimited)",
    )
    parser.add_argument(
        "--type",
        choices=["conversations", "memory", "both"],
        default="both",
        help="What to import (default: both)",
    )
    parser.add_argument(
        "--model",
        choices=["openrouter", "ollama"],
        default="openrouter",
        help="LLM backend (default: openrouter)",
    )
    parser.add_argument(
        "--ollama-model", default="qwen3", help="Ollama model name (default: qwen3)"
    )
    parser.add_argument(
        "--verbose", action="store_true", help="Show full content during processing"
    )
    parser.add_argument(
        "--report",
        type=str,
        metavar="FILE",
        help="Write a markdown report of everything imported",
    )
    return parser.parse_args()


# ─── Main: Conversations Pipeline ───────────────────────────────────────────


def process_conversations(conversations, sync_log, args):
    """Process and ingest conversations. Returns stats dict."""
    stats = {
        "total": len(conversations),
        "already_imported": 0,
        "processed": 0,
        "thoughts_generated": 0,
        "ingested": 0,
        "errors": 0,
        "report_entries": [],
    }

    if not conversations:
        return stats

    print(f"\n{'═' * 60}")
    print("Conversations")
    print(f"{'═' * 60}")

    for conv in conversations:
        if args.limit and stats["processed"] >= args.limit:
            break

        skip_reason = should_skip_conversation(conv, sync_log, args)
        if skip_reason:
            if skip_reason == "already_imported":
                stats["already_imported"] += 1
            continue

        stats["processed"] += 1
        uuid = conv["uuid"]
        title = conv["title"] or "(untitled)"
        answer_text = conv["answer_text"]

        # Parse date
        created = conv.get("created", "")
        created_iso = parse_timestamp_iso(created)
        date_str = created_iso[:10] if created_iso else ""

        word_count = len(answer_text.split())
        print(f"\n{stats['processed']}. {title}")
        print(f"   {word_count} words | {date_str} | {uuid[:8]}...")

        if not answer_text.strip():
            print("   -> No answer text, skipping")
            continue

        # Summarize
        thoughts = summarize(title, date_str, answer_text, args)
        stats["thoughts_generated"] += len(thoughts)

        if not thoughts:
            print("   -> No thoughts extracted (empty summary)")
            if not args.dry_run:
                dedupe_key = make_dedupe_key(uuid)
                sync_log["ingested_ids"][dedupe_key] = datetime.now(
                    timezone.utc
                ).isoformat()
                save_sync_log(sync_log)
            continue

        if args.verbose or args.dry_run:
            for i, thought in enumerate(thoughts, 1):
                preview = thought if len(thought) <= 200 else thought[:200] + "..."
                print(f"   Thought {i}: {preview}")

        if args.report:
            stats["report_entries"].append(
                {
                    "title": title,
                    "date": date_str,
                    "words": word_count,
                    "thoughts": thoughts,
                }
            )

        if args.dry_run:
            continue

        # Build metadata
        metadata = {
            "source": "perplexity",
            "perplexity_title": title,
            "perplexity_date": date_str,
            "perplexity_uuid": uuid,
        }

        # Ingest thoughts
        all_ok = True
        for i, thought in enumerate(thoughts):
            content = f"[Perplexity: {title} | {date_str}] {thought}"
            result = ingest_thought_supabase(content, metadata, created_at=created_iso)

            if result.get("ok"):
                stats["ingested"] += 1
                print(f"   -> Thought {i + 1} ingested")
            else:
                stats["errors"] += 1
                all_ok = False
                print(
                    f"   -> ERROR (thought {i + 1}): {result.get('error', 'unknown')}"
                )

            time.sleep(0.2)

        if all_ok:
            dedupe_key = make_dedupe_key(uuid)
            sync_log["ingested_ids"][dedupe_key] = datetime.now(
                timezone.utc
            ).isoformat()
            save_sync_log(sync_log)

    return stats


# ─── Main: Memory Pipeline ──────────────────────────────────────────────────


def process_memory(memories, sync_log, args):
    """Process and ingest memory rows. Returns stats dict."""
    stats = {
        "total": len(memories),
        "already_imported": 0,
        "deleted": 0,
        "forgotten": 0,
        "processed": 0,
        "thoughts_generated": 0,
        "ingested": 0,
        "errors": 0,
        "report_entries": [],
    }

    if not memories:
        return stats

    print(f"\n{'═' * 60}")
    print("Memory")
    print(f"{'═' * 60}")

    for mem in memories:
        if args.limit and stats["processed"] >= args.limit:
            break

        skip_reason = should_skip_memory(mem, sync_log)
        if skip_reason:
            if skip_reason == "already_imported":
                stats["already_imported"] += 1
            elif skip_reason == "deleted":
                stats["deleted"] += 1
            elif skip_reason == "forgotten":
                stats["forgotten"] += 1
            continue

        stats["processed"] += 1

        # Build items to ingest: list of (dedupe_key, synthetic_key, text, metadata, created_at_iso)
        items = []
        is_profile = is_json_profile_row(mem)

        if is_profile:
            try:
                profile_json = json.loads(mem["MEMORY_VALUE"])
            except (json.JSONDecodeError, TypeError):
                print(f"\n   Warning: Failed to parse JSON profile, skipping")
                stats["processed"] -= 1
                continue

            profile_entries = flatten_json_profile(profile_json)
            print(
                f"\n{stats['processed']}. [JSON Profile] ({len(profile_entries)} sections)"
            )

            first_created = mem.get("FIRST_CREATED_AT", "")
            created_iso = parse_timestamp_iso(first_created)

            for synthetic_key, text in profile_entries:
                dedupe_key = make_dedupe_key("json_profile", text[:200])
                meta = {
                    "source": "perplexity_memory",
                    "memory_key": synthetic_key,
                    "memory_confidence": "high",
                    "memory_first_created": first_created,
                    "memory_profile_section": synthetic_key.split(".", 1)[-1],
                }
                items.append((dedupe_key, synthetic_key, text, meta, created_iso))

        else:
            memory_key = mem.get("MEMORY_KEY", "")
            memory_value = mem.get("MEMORY_VALUE", "")
            confidence = mem.get("CONFIDENCE", "")
            first_created = mem.get("FIRST_CREATED_AT", "")
            source_query = mem.get("LAST_UPDATED_QUERY", "")
            created_iso = parse_timestamp_iso(first_created)

            print(f"\n{stats['processed']}. [{memory_key}]")

            dedupe_key = make_dedupe_key(memory_key, first_created)
            meta = {
                "source": "perplexity_memory",
                "memory_key": memory_key,
                "memory_confidence": confidence,
                "memory_first_created": first_created,
                "memory_source_query": source_query,
            }
            items.append((dedupe_key, memory_key, memory_value, meta, created_iso))

        stats["thoughts_generated"] += len(items)

        if args.verbose or args.dry_run:
            for i, (_, key, text, _, _) in enumerate(items, 1):
                preview = text if len(text) <= 200 else text[:200] + "..."
                print(f"   Thought {i} [{key}]: {preview}")

        if args.report:
            if is_profile:
                stats["report_entries"].append(
                    {
                        "label": "[JSON Profile]",
                        "key": "profile",
                        "thoughts": [text for _, _, text, _, _ in items],
                    }
                )
            else:
                stats["report_entries"].append(
                    {
                        "label": mem.get("MEMORY_KEY", ""),
                        "key": mem.get("MEMORY_KEY", ""),
                        "thoughts": [text for _, _, text, _, _ in items],
                    }
                )

        if args.dry_run:
            continue

        # Ingest each item
        all_ok = True
        for i, (dedupe_key, synthetic_key, text, meta, created_iso) in enumerate(items):
            if is_profile:
                section_label = (
                    synthetic_key.split(".", 1)[-1].replace("_", " ").title()
                )
                content = f"[Perplexity Memory: Profile — {section_label}] {text}"
            else:
                content = f"[Perplexity Memory: {synthetic_key}] {text}"

            result = ingest_thought_supabase(content, meta, created_at=created_iso)

            if result.get("ok"):
                stats["ingested"] += 1
                print(f"   -> Thought {i + 1} ingested")
            else:
                stats["errors"] += 1
                all_ok = False
                print(
                    f"   -> ERROR (thought {i + 1}): {result.get('error', 'unknown')}"
                )

            time.sleep(0.2)

        if all_ok:
            for dedupe_key, _, _, _, _ in items:
                sync_log["ingested_ids"][dedupe_key] = datetime.now(
                    timezone.utc
                ).isoformat()
            save_sync_log(sync_log)

    return stats


# ─── Main Entry Point ───────────────────────────────────────────────────────


def main():
    args = parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.is_file():
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    # Validate env vars for live mode
    if not args.dry_run:
        if not SUPABASE_URL:
            print("Error: SUPABASE_URL environment variable required.")
            print(
                "Set it to your Supabase project URL (e.g., https://xxxxx.supabase.co)"
            )
            sys.exit(1)
        if not SUPABASE_SERVICE_ROLE_KEY:
            print("Error: SUPABASE_SERVICE_ROLE_KEY environment variable required.")
            print(
                "This is your Supabase Secret Key (Settings → API → Secret key, starts with sb_secret_)"
            )
            sys.exit(1)
        if not OPENROUTER_API_KEY:
            print(
                "Error: OPENROUTER_API_KEY required for embeddings and summarization."
            )
            print("Get one at https://openrouter.ai/keys")
            sys.exit(1)

    # Warn about missing API key for summarization in dry-run (won't produce summaries)
    if args.dry_run and args.model == "openrouter" and not OPENROUTER_API_KEY:
        print(
            "Note: OPENROUTER_API_KEY not set. Summarization will be skipped in dry-run."
        )
        print("Set the key for a full dry-run preview, or use --model ollama.\n")

    # Display run configuration
    mode = "DRY RUN" if args.dry_run else "LIVE"
    summarize_mode = f"{args.model}"
    if args.model == "ollama":
        summarize_mode += f" ({args.ollama_model})"
    print(f"\n  Mode:        {mode}")
    print(f"  Summarizer:  {summarize_mode}")
    print(f"  Type:        {args.type}")
    if args.after:
        print(f"  After:       {args.after}")
    if args.before:
        print(f"  Before:      {args.before}")
    if args.limit:
        print(f"  Limit:       {args.limit} per type")
    print()

    sync_log = load_sync_log()

    # Process conversations
    conv_stats = None
    if args.type in ("conversations", "both"):
        print(f"Extracting conversations from {xlsx_path}...")
        conversations = extract_conversations(str(xlsx_path))
        print(f"Found {len(conversations)} conversations.")
        conversations.sort(key=lambda c: c.get("created", ""))
        conv_stats = process_conversations(conversations, sync_log, args)

    # Process memory
    mem_stats = None
    if args.type in ("memory", "both"):
        print(f"\nExtracting memory from {xlsx_path}...")
        memories = extract_memory_rows(str(xlsx_path))
        print(f"Found {len(memories)} memory entries.")
        mem_stats = process_memory(memories, sync_log, args)

    # ─── Summary ─────────────────────────────────────────────────────────────

    print(f"\n{'─' * 60}")
    print("Summary:")

    if conv_stats:
        print(f"\n  Conversations:")
        print(f"    Found:              {conv_stats['total']}")
        if conv_stats["already_imported"]:
            print(f"    Already imported:   {conv_stats['already_imported']} (skipped)")
        print(f"    Processed:          {conv_stats['processed']}")
        print(f"    Thoughts:           {conv_stats['thoughts_generated']}")
        if not args.dry_run:
            print(f"    Ingested:           {conv_stats['ingested']}")
            print(f"    Errors:             {conv_stats['errors']}")

    if mem_stats:
        print(f"\n  Memory:")
        print(f"    Found:              {mem_stats['total']}")
        if mem_stats["already_imported"]:
            print(f"    Already imported:   {mem_stats['already_imported']} (skipped)")
        if mem_stats["deleted"]:
            print(f"    Deleted:            {mem_stats['deleted']} (skipped)")
        if mem_stats["forgotten"]:
            print(f"    Forgotten:          {mem_stats['forgotten']} (skipped)")
        print(f"    Processed:          {mem_stats['processed']}")
        print(f"    Thoughts:           {mem_stats['thoughts_generated']}")
        if not args.dry_run:
            print(f"    Ingested:           {mem_stats['ingested']}")
            print(f"    Errors:             {mem_stats['errors']}")

    # Cost estimation
    total_thoughts = 0
    total_processed = 0
    if conv_stats:
        total_thoughts += conv_stats["thoughts_generated"]
        total_processed += conv_stats["processed"]
    if mem_stats:
        total_thoughts += mem_stats["thoughts_generated"]
        total_processed += mem_stats["processed"]

    if total_thoughts > 0:
        # Summarization cost (conversations only): gpt-4o-mini via OpenRouter
        # ~$0.15/1M input, ~$0.60/1M output, ~800 tokens in / 200 tokens out per conv
        conv_count = conv_stats["processed"] if conv_stats else 0
        summarize_cost = (conv_count * 800 * 0.15 / 1_000_000) + (
            conv_count * 200 * 0.60 / 1_000_000
        )

        # Embedding cost: $0.02/1M tokens, ~100 tokens per thought
        embedding_cost = total_thoughts * 100 * 0.02 / 1_000_000

        total_cost = summarize_cost + embedding_cost
        print(f"\n  Est. API cost:          ${total_cost:.4f}")
        if conv_count > 0:
            print(f"    Summarization:        ${summarize_cost:.4f}")
        print(f"    Embeddings:           ${embedding_cost:.4f}")

    print(f"{'─' * 60}")

    # Write report
    if args.report:
        _write_report(args.report, conv_stats, mem_stats, args.dry_run)


def _write_report(filepath, conv_stats, mem_stats, dry_run):
    """Write a markdown report of imported data."""
    with open(filepath, "w") as f:
        mode_str = "DRY RUN" if dry_run else "LIVE"
        f.write(f"# Perplexity Import Report ({mode_str})\n\n")
        f.write(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
        )

        f.write("## Stats\n\n")
        f.write(f"| Metric | Conversations | Memory |\n")
        f.write(f"|--------|--------------|--------|\n")

        conv_total = conv_stats["total"] if conv_stats else 0
        mem_total = mem_stats["total"] if mem_stats else 0
        conv_already = conv_stats["already_imported"] if conv_stats else 0
        mem_already = mem_stats["already_imported"] if mem_stats else 0
        conv_proc = conv_stats["processed"] if conv_stats else 0
        mem_proc = mem_stats["processed"] if mem_stats else 0
        conv_thoughts = conv_stats["thoughts_generated"] if conv_stats else 0
        mem_thoughts = mem_stats["thoughts_generated"] if mem_stats else 0
        conv_ingested = conv_stats["ingested"] if conv_stats else 0
        mem_ingested = mem_stats["ingested"] if mem_stats else 0
        conv_errors = conv_stats["errors"] if conv_stats else 0
        mem_errors = mem_stats["errors"] if mem_stats else 0

        f.write(f"| Found | {conv_total} | {mem_total} |\n")
        f.write(f"| Already imported | {conv_already} | {mem_already} |\n")
        f.write(f"| Processed | {conv_proc} | {mem_proc} |\n")
        f.write(f"| Thoughts | {conv_thoughts} | {mem_thoughts} |\n")
        if not dry_run:
            f.write(f"| Ingested | {conv_ingested} | {mem_ingested} |\n")
            f.write(f"| Errors | {conv_errors} | {mem_errors} |\n")
        f.write("\n")

        # Conversation details
        if conv_stats and conv_stats.get("report_entries"):
            f.write("## Conversations\n\n")
            for entry in conv_stats["report_entries"]:
                f.write(f"### {entry['title']} ({entry['date']})\n\n")
                f.write(f"_{entry['words']} words_\n\n")
                for i, thought in enumerate(entry["thoughts"], 1):
                    f.write(f"{i}. {thought}\n")
                f.write("\n")

        # Memory details
        if mem_stats and mem_stats.get("report_entries"):
            f.write("## Memory\n\n")
            for entry in mem_stats["report_entries"]:
                f.write(f"### {entry['label']}\n\n")
                for i, thought in enumerate(entry["thoughts"], 1):
                    f.write(f"{i}. {thought}\n")
                f.write("\n")

    print(f"\nReport written to {filepath}")


if __name__ == "__main__":
    main()
